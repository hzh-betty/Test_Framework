import os
from openpyxl import load_workbook
from openpyxl.styles import Font
from conf import setting
from common.record_log import logs

class OperationExcel:
    """Class for handling Excel file operations."""

    def __init__(self, file_path=None):
        self.__file_path = file_path if file_path else setting.FILE_PATH['EXCEL']
        self.__sheet_id = setting.SHEET_ID
        self.__GLOBAL_TABLE  = self._load_workbook()

    def _load_workbook(self):
        """Load the Excel workbook."""
        if os.path.splitext(self.__file_path)[-1] != '.xlsx':
            logs.error(f"Unsupported file format: {self.__file_path}")
            raise ValueError("Only .xlsx files are supported.")
        wb = load_workbook(self.__file_path)
        sheet_names = wb.sheetnames

        return wb[sheet_names[self.__sheet_id]]

    def get_rows(self):
        """Get the number of rows in the sheet."""
        return self.__GLOBAL_TABLE.max_row

    def get_cols(self):
        """Get the number of columns in the sheet."""
        return self.__GLOBAL_TABLE.max_column

    def get_cell_value(self, row, col):
        """Get the value of a specific cell."""
        return self.__GLOBAL_TABLE.cell(row=row, column=col).value

    def write_value(self, row, col, value):
        """Write a value to a specific cell."""
        try:
            self.__GLOBAL_TABLE.cell(row=row, column=col, value=value)
            self.__GLOBAL_TABLE.parent.save(self.__file_path)
        except PermissionError:
            logs.error("Failed to write to Excel file. Please close the file if it's open.")
            return

    def get_each_line(self, row):
        """ get entire row """
        return [cell.value for cell in list(self.__GLOBAL_TABLE[row])]

    def get_each_column(self, col=None):
        """ get entire column """
        col = col if col else 1
        return [self.__GLOBAL_TABLE.cell(row=i, column=col).value for i in range(1, self.get_rows() + 1)]

    def setting_style(self, row, col, bold=True, color="FF0000"):
        """ Set the font style of a specific cell. """
        cell = self.__GLOBAL_TABLE.cell(row=row, column=col)
        cell.font = Font(bold=bold, color=color)
        self.__GLOBAL_TABLE.parent.save(self.__file_path)

if __name__ == "__main__":
    excel = OperationExcel()
    print(excel.get_cell_value(1, 2))
    excel.write_value(1, 2, "Test Value")
    print(excel.get_each_line(1))
    print(excel.get_each_column(2))
    # excel.setting_style(1, 2, bold=True, color="00FF00")
    excel.write_value(1, 2, "password")